# processos/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Processo, ProcessHistory, MonitoramentoRecord, Profile
import json
from datetime import datetime, date, timedelta, time
from django.db.models import Q
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import CustomUserCreationForm
from django.contrib.auth.models import User
from .forms import ProcessoForm
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from django.utils import timezone, dateformat
from django.contrib.auth.forms import AuthenticationForm


# Helper function to check user permissions based on level
def can_access_genero(user, genero):
    if user.is_superuser:  # Superusers can access all
        return True

    try:
        user_level = user.profile.level
    except Profile.DoesNotExist:
        return False  # No profile, no access

    if user_level == '3':  # Usuário Geral can access all
        return True
    elif user_level == '1' and genero == 'LICITACOES_E_CONTRATOS':  # Analista 1 only Licitações
        return True
    elif user_level == '2' and genero == 'LIQUIDACOES':  # Analista 2 only Liquidações
        return True
    elif user_level == '0':  # Protocolo has no specific genre restrictions for general views, but specific functions
        return True
    return False

# Helper function to filter querysets by user level


def filter_processes_by_user_level(user, queryset):
    if user.is_superuser:
        return queryset  # Superusers see all

    try:
        user_level = user.profile.level
    except Profile.DoesNotExist:
        return queryset.none()  # If no profile, no processes

    if user_level == '3':  # Usuário Geral sees all
        return queryset
    elif user_level == '1':  # Analista 1 sees only LICITACOES_E_CONTRATOS
        return queryset.filter(genero='LICITACOES_E_CONTRATOS')
    elif user_level == '2':  # Analista 2 sees only LIQUIDACOES
        return queryset.filter(genero='LIQUIDACOES')
    elif user_level == '0':  # Protocolo sees all processes, but might have limitations on functions
        return queryset
    return queryset.none()  # Default to no access if level is not recognized


@login_required
def cadastrar_processo(request):
    # Only Protocolo (0) and Usuário Geral (3) can create new processes
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.level in ['0', '3'])):
        return HttpResponse("Você não tem permissão para cadastrar novos processos.", status=403)

    if request.method == 'POST':
        form = ProcessoForm(request.POST)
        if form.is_valid():
            processo = form.save(commit=False)

            # Ensure the user has permission to create this type of process
            if not can_access_genero(request.user, processo.genero):
                return HttpResponse("Você não tem permissão para cadastrar processos deste gênero.", status=403)

            if processo.prioridade == 'SIM':
                processo.prazo_dias = 2
            else:
                processo.prazo_dias = 7

            processo.save()
            return redirect('listar_processos')
    else:
        form = ProcessoForm()
    return render(request, 'Formulario.html', {'form': form})


def calcular_prazo(data_entrada, prioridade_value):
    if not data_entrada or not prioridade_value:
        return None
    try:
        if prioridade_value == 'SIM':
            dias_prazo = 2
        elif prioridade_value == 'NAO':
            dias_prazo = 7
        else:
            dias_prazo = 7

        data_prazo = data_entrada + timedelta(days=dias_prazo)
        return data_prazo
    except Exception as e:
        print(f"Erro ao calcular prazo: {e}")
        return None


def calcular_proxima_data_monitoramento(data_base, prazo_monitoramento_tipo):
    """
    Calcula a próxima data de monitoramento com base na data base e no tipo de prazo.
    """
    if not data_base or not prazo_monitoramento_tipo:
        return None

    if prazo_monitoramento_tipo == 'SEMESTRAL':
        return data_base + timedelta(days=6*30)  # Aproximadamente 6 meses
    elif prazo_monitoramento_tipo == 'QUADRIMESTRAL':
        return data_base + timedelta(days=4*30)  # Aproximadamente 4 meses
    elif prazo_monitoramento_tipo == 'ANUAL':
        return data_base + timedelta(days=12*30)  # Aproximadamente 12 meses
    elif prazo_monitoramento_tipo == 'TRIMESTRAL':
        return data_base + timedelta(days=3*30)  # Aproximadamente 3 meses
    return None


@login_required
def index(request):
    return render(request, 'Formulario.html')


@csrf_exempt
@login_required
def salvar_processo(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            data_entrada_obj = datetime.strptime(
                data['data_entrada'], '%Y-%m-%d').date()
            hora_entrada_obj = datetime.strptime(
                data['hora_entrada'], '%H:%M').time()
            data_saida_obj = datetime.strptime(
                data['data_saida'], '%Y-%m-%d').date() if data.get('data_saida') else None
            hora_saida_obj = datetime.strptime(
                data['hora_saida'], '%H:%M').time() if data.get('hora_saida') else None

            # Permission check: Only Protocolo (0) and Usuário Geral (3) can save new processes
            if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.level in ['0', '3'])):
                return JsonResponse({"success": False, "message": "Você não tem permissão para salvar novos processos."}, status=403)

            if not can_access_genero(request.user, data['genero']):
                return JsonResponse({"success": False, "message": "Você não tem permissão para salvar processos deste gênero."}, status=403)

            prazo_dias_value = None
            if data['prioridade'] == 'SIM':
                prazo_dias_value = 2
            elif data['prioridade'] == 'NAO':
                prazo_dias_value = 7

            # Lógica de definição do prazo de monitoramento
            prazo_monitoramento_value = 'NAO_APLICAVEL'
            proxima_data_monitoramento_value = None
            status_monitoramento_value = 'NAO_APLICAVEL'

            especie_processo = data.get('especie')
            genero_processo = data.get('genero')

            # Determina a data base para o cálculo do monitoramento.
            # Para tipos de monitoramento que iniciam após uma "finalização" (como P.C. ou Concessão),
            # a data_saida é o ponto de partida lógico. Caso contrário, data_entrada.
            data_base_monitoramento = data_saida_obj if data_saida_obj else data_entrada_obj

            # Apply monitoring logic based on genre and species
            if genero_processo == 'LIQUIDACOES':
                if especie_processo == 'P.C. Bolsa Atleta':
                    prazo_monitoramento_value = 'SEMESTRAL'
                    status_monitoramento_value = 'PENDENTE'
                elif especie_processo == 'P.C. Adiantamento':
                    prazo_monitoramento_value = 'TRIMESTRAL'
                    status_monitoramento_value = 'PENDENTE'
                elif especie_processo == 'Subvenção Social - Prestação de Contas':
                    prazo_monitoramento_value = 'QUADRIMESTRAL'
                    status_monitoramento_value = 'PENDENTE'
                elif especie_processo == 'Subvenção Social - P.C. Anual':
                    prazo_monitoramento_value = 'ANUAL'
                    status_monitoramento_value = 'PENDENTE'
                elif especie_processo == 'P.C. Subvenção Bloco Carnaval':
                    prazo_monitoramento_value = 'TRIMESTRAL'
                    status_monitoramento_value = 'PENDENTE'
                elif especie_processo == 'P.C. Patrocínio':
                    prazo_monitoramento_value = 'TRIMESTRAL'
                    status_monitoramento_value = 'PENDENTE'
                elif especie_processo in ['Concessão Aux. Bolsa Atleta', 'Concessão Aux. Aluguel Social',
                                          'Concessão Adiantamento', 'Subvenção Social - Concessão', 'Concessão Diária']:
                    prazo_monitoramento_value = 'TRIMESTRAL'
                    status_monitoramento_value = 'PENDENTE'
            elif genero_processo == 'LICITACOES_E_CONTRATOS':
                # Only 'Concessão Patrocínio' needs monitoring in this genre
                if especie_processo == 'Concessão Patrocínio':
                    prazo_monitoramento_value = 'TRIMESTRAL'
                    status_monitoramento_value = 'PENDENTE'
                else:
                    # Other species in LICITACOES_E_CONTRATOS do not need monitoring
                    prazo_monitoramento_value = 'NAO_APLICAVEL'
                    status_monitoramento_value = 'NAO_APLICAVEL'
            else:  # For 'OUTROS_GENERO' or any other unlisted genre
                prazo_monitoramento_value = 'NAO_APLICAVEL'
                status_monitoramento_value = 'NAO_APLICAVEL'

            if prazo_monitoramento_value != 'NAO_APLICAVEL' and data_base_monitoramento:
                proxima_data_monitoramento_value = calcular_proxima_data_monitoramento(
                    data_base_monitoramento, prazo_monitoramento_value)

            processo = Processo.objects.create(
                numero_processo=data['numero_processo'],
                volume=data.get('volume') or None,
                secretaria=data['secretaria'],
                data_entrada=data_entrada_obj,
                hora_entrada=hora_entrada_obj,
                data_saida=data_saida_obj,
                hora_saida=hora_saida_obj,
                destino=data.get('destino') or None,
                genero=genero_processo,  # Use the determined genre
                especie=especie_processo,  # Use the determined species
                objeto=data['objeto'],
                contratada=data.get('contratada') or None,
                recorrente=data.get('recorrente', 'NAO'),
                prioridade=data['prioridade'],
                prazo_dias=prazo_dias_value,
                tecnico=data.get('tecnico') or None,
                numero_despacho=data.get('numero_despacho') or None,
                observacao=data.get('observacao') or None,
                prazo_monitoramento=prazo_monitoramento_value,
                proxima_data_monitoramento=proxima_data_monitoramento_value,
                status_monitoramento=status_monitoramento_value,
                # NEW FIELDS
                valor=data.get('valor') or None,
                periodo=data.get('periodo') or None,
                status_analise=data.get('status_analise', 'NAO_APLICAVEL')
            )

            # Lógica para concluir monitoramento de processos anteriores com o mesmo número
            # e que são da espécie 'Concessão' (ou similar) que indicam uma nova PC sendo enviada
            especies_para_finalizar_automaticamente = [
                'P.C. Bolsa Atleta', 'Subvenção Social - Prestação de Contas',
                'Subvenção Social - P.C. Anual', 'Concessão Patrocínio',
                'P.C. Adiantamento', 'P.C. Patrocínio', 'P.C. Subvenção Bloco Carnaval',
                'Concessão Aux. Bolsa Atleta', 'Concessão Aux. Aluguel Social',
                'Concessão Adiantamento', 'Subvenção Social - Concessão', 'Concessão Diária'
            ]

            # Só dispara se o NOVO processo for um que indica que uma PC está sendo submetida
            if especie_processo in especies_para_finalizar_automaticamente:
                processos_anteriores = Processo.objects.filter(
                    numero_processo=data['numero_processo'],
                    # Busca por processos anteriores desses tipos específicos
                    especie__in=especies_para_finalizar_automaticamente
                ).exclude(id=processo.id)  # Exclui o processo que acabou de ser criado

                for p_antigo in processos_anteriores:
                    if p_antigo.status_monitoramento in ['PENDENTE', 'ATRASADO']:
                        p_antigo.status_monitoramento = 'CONCLUIDO'
                        # Adiciona um registro de monitoramento
                        MonitoramentoRecord.objects.create(
                            processo=p_antigo,
                            data_registro=date.today(),
                            observacao=f"Monitoramento concluído automaticamente pela entrada de um novo processo com o mesmo número ({processo.numero_processo}) e espécie relacionada: {especie_processo}.",
                            registrado_por=request.user
                        )
                        # Remove a próxima data de monitoramento e o tipo de prazo para "desativar" monitoramento para o ciclo anterior.
                        p_antigo.proxima_data_monitoramento = None
                        p_antigo.prazo_monitoramento = 'NAO_APLICAVEL'
                        p_antigo.save()

            return JsonResponse({"success": True, "message": "Processo salvo!", "id": processo.id})
        except Exception as e:
            print(f"Erro em salvar_processo: {e}")
            return JsonResponse({"success": False, "message": str(e)}, status=400)
    return JsonResponse({"success": False, "message": "Método não permitido"}, status=405)


@login_required
def listar_processos(request):
    termo_pesquisa = request.GET.get('termo', '').strip()
    prioridade_filtro = request.GET.get('prioridade', 'todas')
    genero_filtro = request.GET.get('genero', 'todas')
    especie_filtro = request.GET.get('especie', 'todas')

    base_query = Processo.objects.filter(data_saida__isnull=True)
    processos_query = filter_processes_by_user_level(request.user, base_query)

    if termo_pesquisa:
        processos_query = processos_query.filter(
            Q(numero_processo__icontains=termo_pesquisa) |
            Q(secretaria__icontains=termo_pesquisa) |
            Q(objeto__icontains=termo_pesquisa) |
            Q(contratada__icontains=termo_pesquisa) |
            Q(tecnico__icontains=termo_pesquisa) |
            Q(valor__icontains=termo_pesquisa) |
            Q(periodo__icontains=termo_pesquisa)
        )

    if prioridade_filtro != 'todas':
        processos_query = processos_query.filter(prioridade=prioridade_filtro)

    if genero_filtro != 'todas':
        if request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.level == '3'):
            processos_query = processos_query.filter(genero=genero_filtro)
        elif (hasattr(request.user, 'profile') and request.user.profile.level == '1' and genero_filtro == 'LICITACOES_E_CONTRATOS'):
            processos_query = processos_query.filter(
                genero='LICITACOES_E_CONTRATOS')
        elif (hasattr(request.user, 'profile') and request.user.profile.level == '2' and genero_filtro == 'LIQUIDACOES'):
            processos_query = processos_query.filter(genero='LIQUIDACOES')
        else:
            processos_query = processos_query.none()

    if especie_filtro != 'todas':
        processos_query = processos_query.filter(especie=especie_filtro)

    processos_query = processos_query.order_by('data_entrada', 'prioridade')

    for processo in processos_query:
        prazo_obj = calcular_prazo(processo.data_entrada, processo.prioridade)
        if prazo_obj:
            dias_restantes = (prazo_obj - date.today()).days
            processo.prazo_formatado = f"{dias_restantes} dia(s) {'atrasado' if dias_restantes < 0 else 'restante(s)'}"
        else:
            processo.prazo_formatado = "-"

    all_generos = Processo.objects.values_list(
        'genero', flat=True).distinct().order_by('genero')
    all_generos = [
        g for g in all_generos if can_access_genero(request.user, g)]

    all_especies = []
    if genero_filtro != 'todas':
        if can_access_genero(request.user, genero_filtro):
            all_especies = Processo.objects.filter(genero=genero_filtro).values_list(
                'especie', flat=True).distinct().order_by('especie')
    else:
        allowed_generos_for_all_species = [g for g in Processo.objects.values_list(
            'genero', flat=True).distinct() if can_access_genero(request.user, g)]
        all_especies = Processo.objects.filter(genero__in=allowed_generos_for_all_species).values_list(
            'especie', flat=True).distinct().order_by('especie')

    # Pre-calculate flags for template logic
    can_edit = request.user.is_superuser or \
        (hasattr(request.user, 'profile')
         and request.user.profile.level in ['1', '2', '3'])

    can_delete = request.user.is_superuser

    can_mark_saida = request.user.is_superuser or \
        (hasattr(request.user, 'profile')
         and request.user.profile.level in ['0', '3'])

    can_create_process = request.user.is_superuser or \
        (hasattr(request.user, 'profile')
         and request.user.profile.level in ['0', '3'])

    return render(request, 'lista_processos.html', {
        'processos': processos_query,
        'prioridade_filtro': prioridade_filtro,
        'termo_pesquisa': termo_pesquisa,
        'genero_filtro': genero_filtro,
        'especie_filtro': especie_filtro,
        'all_generos': all_generos,
        'all_especies': all_especies,
        'user_level': request.user.profile.level if hasattr(request.user, 'profile') else None,
        'can_edit': can_edit,
        'can_delete': can_delete,
        'can_mark_saida': can_mark_saida,
        'can_create_process': can_create_process,
    })


@csrf_exempt
@login_required
def atualizar_processo(request, id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            processo = get_object_or_404(Processo, id=id)

            # Permission check for updating a process
            if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.level in ['1', '2', '3'])):
                return JsonResponse({"success": False, "message": "Você não tem permissão para editar processos."}, status=403)

            if not can_access_genero(request.user, processo.genero):
                return JsonResponse({"success": False, "message": "Você não tem permissão para editar este processo."}, status=403)

            if 'genero' in data and data['genero'] != processo.genero:
                if not can_access_genero(request.user, data['genero']):
                    return JsonResponse({"success": False, "message": "Você não tem permissão para alterar o gênero para este valor."}, status=403)

            changed_fields = {}
            original_values = {field.name: getattr(
                processo, field.name) for field in Processo._meta.fields}

            for field_name, new_value in data.items():
                current_value = getattr(processo, field_name)

                converted_value = None
                if field_name.startswith('data_') and new_value:
                    converted_value = datetime.strptime(
                        new_value, '%Y-%m-%d').date()
                elif field_name.startswith('hora_') and new_value:
                    converted_value = datetime.strptime(
                        new_value, '%H:%M').time()
                elif new_value == '':
                    converted_value = None
                else:
                    converted_value = new_value

                # Special handling for 'status_analise' to use default if not provided
                if field_name == 'status_analise' and (new_value == '' or new_value is None):
                    converted_value = 'NAO_APLICAVEL'

                if str(current_value) != str(converted_value):
                    setattr(processo, field_name, converted_value)

                    old_val_str = ''
                    if original_values[field_name] is not None:
                        if isinstance(original_values[field_name], date):
                            old_val_str = original_values[field_name].strftime(
                                '%Y-%m-%d')
                        elif isinstance(original_values[field_name], time):
                            old_val_str = original_values[field_name].strftime(
                                '%H:%M')
                        elif field_name == 'status_analise':
                            # Get display value for old status_analise
                            old_val_str = dict(Processo.STATUS_ANALISE_CHOICES).get(
                                original_values[field_name], original_values[field_name])
                        else:
                            old_val_str = str(original_values[field_name])

                    new_val_str = ''
                    if converted_value is not None:
                        if isinstance(converted_value, date):
                            new_val_str = converted_value.strftime('%Y-%m-%d')
                        elif isinstance(converted_value, time):
                            new_val_str = converted_value.strftime('%H:%M')
                        elif field_name == 'status_analise':
                            # Get display value for new status_analise
                            new_val_str = dict(Processo.STATUS_ANALISE_CHOICES).get(
                                converted_value, converted_value)
                        else:
                            new_val_str = str(converted_value)

                    changed_fields[field_name] = {
                        'old': old_val_str, 'new': new_val_str}

            if 'prioridade' in data:
                if data['prioridade'] == 'SIM':
                    processo.prazo_dias = 2
                elif data['prioridade'] == 'NAO':
                    processo.prazo_dias = 7

            data_entrada_actual = processo.data_entrada
            data_saida_actual = processo.data_saida
            especie_actual = processo.especie
            genero_actual = processo.genero

            data_base_monitoramento_recalc = data_saida_actual if data_saida_actual else data_entrada_actual

            novo_prazo_monitoramento_value = 'NAO_APLICAVEL'
            novo_status_monitoramento_value = 'NAO_APLICAVEL'

            # Re-evaluate monitoring based on updated genre and species
            if genero_actual == 'LIQUIDACOES':
                if especie_actual == 'P.C. Bolsa Atleta':
                    novo_prazo_monitoramento_value = 'SEMESTRAL'
                    novo_status_monitoramento_value = 'PENDENTE'
                elif especie_actual == 'P.C. Adiantamento':
                    novo_prazo_monitoramento_value = 'TRIMESTRAL'
                    novo_status_monitoramento_value = 'PENDENTE'
                elif especie_actual == 'Subvenção Social - Prestação de Contas':
                    novo_prazo_monitoramento_value = 'QUADRIMESTRAL'
                    novo_status_monitoramento_value = 'PENDENTE'
                elif especie_actual == 'Subvenção Social - P.C. Anual':
                    novo_prazo_monitoramento_value = 'ANUAL'
                    novo_status_monitoramento_value = 'PENDENTE'
                elif especie_actual == 'Concessão Patrocínio':
                    novo_prazo_monitoramento_value = 'TRIMESTRAL'
                    novo_status_monitoramento_value = 'PENDENTE'
                elif especie_actual == 'P.C. Subvenção Bloco Carnaval':
                    novo_prazo_monitoramento_value = 'TRIMESTRAL'
                    novo_status_monitoramento_value = 'PENDENTE'
                elif especie_actual == 'P.C. Patrocínio':
                    novo_prazo_monitoramento_value = 'TRIMESTRAL'
                    novo_status_monitoramento_value = 'PENDENTE'
                elif especie_actual in ['Concessão Aux. Bolsa Atleta', 'Concessão Aux. Aluguel Social',
                                        'Concessão Adiantamento', 'Subvenção Social - Concessão', 'Concessão Diária']:
                    novo_prazo_monitoramento_value = 'TRIMESTRAL'
                    novo_status_monitoramento_value = 'PENDENTE'
            elif genero_actual == 'LICITACOES_E_CONTRATOS':

                if especie_actual == 'Concessão Patrocínio':
                    novo_prazo_monitoramento_value = 'TRIMESTRAL'
                    novo_status_monitoramento_value = 'PENDENTE'
                else:
                    # Other species in LICITACOES_E_CONTRATOS do not need monitoring
                    novo_prazo_monitoramento_value = 'NAO_APLICAVEL'
                    novo_status_monitoramento_value = 'NAO_APLICAVEL'
            else:
                # For 'OUTROS_GENERO' or any other unlisted genre
                novo_prazo_monitoramento_value = 'NAO_APLICAVEL'
                novo_status_monitoramento_value = 'NAO_APLICAVEL'

            if (processo.prazo_monitoramento != novo_prazo_monitoramento_value) or \
               (processo.status_monitoramento == 'NAO_APLICAVEL' and novo_status_monitoramento_value == 'PENDENTE') or \
               ('data_saida' in changed_fields and data_saida_actual is not None and novo_prazo_monitoramento_value != 'NAO_APLICAVEL') or \
               ('data_entrada' in changed_fields and data_saida_actual is None and novo_prazo_monitoramento_value != 'NAO_APLICAVEL'):

                processo.prazo_monitoramento = novo_prazo_monitoramento_value
                processo.status_monitoramento = novo_status_monitoramento_value

                if processo.prazo_monitoramento != 'NAO_APLICAVEL' and data_base_monitoramento_recalc:
                    processo.proxima_data_monitoramento = calcular_proxima_data_monitoramento(
                        data_base_monitoramento_recalc, processo.prazo_monitoramento)
                else:
                    processo.proxima_data_monitoramento = None
            processo.save()

            for field_name, values in changed_fields.items():
                ProcessHistory.objects.create(
                    process=processo,
                    field_name=field_name,
                    old_value=values['old'],
                    new_value=values['new'],
                    changed_by=request.user
                )
            return JsonResponse({"success": True, "message": "Processo atualizado!"})
        except Processo.DoesNotExist:
            return JsonResponse({"success": False, "message": "Processo não encontrado."}, status=404)
        except Exception as e:
            print(f"Erro em atualizar_processo: {e}")
            return JsonResponse({"success": False, "message": str(e)}, status=400)
    return JsonResponse({"success": False, "message": "Método não permitido"}, status=405)


@csrf_exempt
@login_required
@user_passes_test(lambda u: u.is_superuser)  # Only superusers can delete
def deletar_processo(request, id):
    if request.method == 'POST':
        try:
            processo = get_object_or_404(Processo, id=id)
            processo.delete()
            return JsonResponse({'success': True, 'message': 'Processo deletado com sucesso!'})
        except Processo.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Processo não encontrado.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    return JsonResponse({'success': False, 'message': 'Método não permitido.'}, status=405)


@login_required
def ver_historico_processo(request, process_id):
    processo = get_object_or_404(Processo, id=process_id)
    if not can_access_genero(request.user, processo.genero):
        return HttpResponse("Você não tem permissão para visualizar o histórico deste processo.", status=403)

    history_records = ProcessHistory.objects.filter(
        process=processo).order_by('-changed_at')
    monitoramento_records = MonitoramentoRecord.objects.filter(
        processo=processo).order_by('-data_registro')

    formatted_history = []
    for record in history_records:
        print(
            f"Raw changed_at: {record.changed_at}, Type: {type(record.changed_at)}")
        rec_dict = {
            'field_name': record.field_name,
            'old_value': record.old_value,
            'new_value': record.new_value,
            'changed_at_formatted': dateformat.format(timezone.localtime(record.changed_at), 'd/m/Y H:i:s',),
            'changed_by_user': record.changed_by.username if record.changed_by else 'N/A'
        }
        formatted_history.append(rec_dict)

    formatted_monitoramento_records = []
    for record in monitoramento_records:
        mon_dict = {
            'data_registro': record.data_registro.strftime('%d/%m/%Y %H:%M:%S'),
            'observacao': record.observacao,
            'registrado_por': record.registrado_por.username if record.registrado_por else 'N/A'
        }
        formatted_monitoramento_records.append(mon_dict)

    return render(request, 'historico_processo.html', {
        'history': formatted_history,
        'monitoramento_records': formatted_monitoramento_records,
        'process_id': process_id,
        'process_number': processo.numero_processo,
    })


@csrf_exempt
@login_required
def concluir_monitoramento(request, process_id):
    processo = get_object_or_404(Processo, id=process_id)

    # Permission check for concluding monitoring
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.level in ['1', '2', '3'])):
        return JsonResponse({"success": False, "message": "Você não tem permissão para concluir o monitoramento deste processo."}, status=403)

    if not can_access_genero(request.user, processo.genero):
        return JsonResponse({"success": False, "message": "Você não tem permissão para concluir o monitoramento deste processo."}, status=403)

    # REMOVIDO: A verificação se o status é 'PENDENTE' ou 'ATRASADO'.
    # Agora, o botão pode ser clicado independente do status atual do monitoramento.

    # Obter a data e hora atuais no fuso horário local configurado no settings.py
    now_local = timezone.localtime(timezone.now())

    # Se o processo não tiver data de saída, preenchemos para "finalizá-lo"
    if not processo.data_saida:
        processo.data_saida = now_local.date()
        processo.hora_saida = now_local.time()
        # Registrar no histórico que a data/hora de saída foi marcada
        ProcessHistory.objects.create(
            process=processo,
            field_name='data_saida',
            old_value=None,
            new_value=processo.data_saida.strftime('%Y-%m-%d'),
            changed_by=request.user
        )
        ProcessHistory.objects.create(
            process=processo,
            field_name='hora_saida',
            old_value=None,
            new_value=processo.hora_saida.strftime('%H:%M:%S'),
            changed_by=request.user
        )

    # Sempre atualiza o status do monitoramento para CONCLUIDO e zera a próxima data
    processo.status_monitoramento = 'CONCLUIDO'
    # Zera para evitar futuros monitoramentos
    processo.proxima_data_monitoramento = None
    # Desativa o tipo de prazo de monitoramento
    processo.prazo_monitoramento = 'NAO_APLICAVEL'

    # Adiciona um registro no histórico de monitoramento
    MonitoramentoRecord.objects.create(
        processo=processo,
        data_registro=date.today(),
        observacao=f"Monitoramento concluído manualmente por {request.user.username}.",
        registrado_por=request.user
    )

    processo.save()

    return JsonResponse({
        'success': True,
        'message': 'Monitoramento concluído com sucesso e desativado para este processo.'
    })


@login_required
def listar_finalizados(request):
    data_inicial_filtro = request.GET.get('data_inicial', '').strip()
    data_final_filtro = request.GET.get('data_final', '').strip()
    prioridade_filtro = request.GET.get('prioridade', 'todas')
    termo_pesquisa = request.GET.get('termo', '').strip()
    status_monitoramento_filtro = request.GET.get(
        'status_monitoramento', 'todas')
    especie_filtro = request.GET.get('especie', 'todas')
    genero_filtro = request.GET.get('genero', 'todas')
    status_analise_filtro = request.GET.get('status_analise', 'todas')

    base_query = Processo.objects.exclude(data_saida__isnull=True)
    processos_query = filter_processes_by_user_level(request.user, base_query)

    if termo_pesquisa:
        processos_query = processos_query.filter(
            Q(numero_processo__icontains=termo_pesquisa) |
            Q(secretaria__icontains=termo_pesquisa) |
            Q(objeto__icontains=termo_pesquisa) |
            Q(contratada__icontains=termo_pesquisa) |
            Q(tecnico__icontains=termo_pesquisa) |
            Q(valor__icontains=termo_pesquisa) |
            Q(periodo__icontains=termo_pesquisa)
        )

    if prioridade_filtro != 'todas':
        processos_query = processos_query.filter(prioridade=prioridade_filtro)

    if data_inicial_filtro and data_final_filtro:
        try:
            start_date = datetime.strptime(
                data_inicial_filtro, '%Y-%m-%d').date()
            end_date = datetime.strptime(data_final_filtro, '%Y-%m-%d').date()
            processos_query = processos_query.filter(
                data_saida__range=[start_date, end_date])
        except ValueError:
            pass
    elif data_inicial_filtro:
        try:
            start_date = datetime.strptime(
                data_inicial_filtro, '%Y-%m-%d').date()
            processos_query = processos_query.filter(
                data_saida__gte=start_date)
        except ValueError:
            pass
    elif data_final_filtro:
        try:
            end_date = datetime.strptime(data_final_filtro, '%Y-%m-%d').date()
            processos_query = processos_query.filter(data_saida__lte=end_date)
        except ValueError:
            pass

    if termo_pesquisa:
        processos_query = processos_query.filter(
            Q(numero_processo__icontains=termo_pesquisa) |
            Q(secretaria__icontains=termo_pesquisa) |
            Q(objeto__icontains=termo_pesquisa) |
            Q(contratada__icontains=termo_pesquisa) |
            Q(tecnico__icontains=termo_pesquisa) |
            Q(valor__icontains=termo_pesquisa) |
            Q(periodo__icontains=termo_pesquisa)
        )

    if status_monitoramento_filtro != 'todas':
        processos_query = processos_query.filter(
            status_monitoramento=status_monitoramento_filtro)

    if genero_filtro != 'todas':
        if request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.level == '3'):
            processos_query = processos_query.filter(genero=genero_filtro)
        elif (hasattr(request.user, 'profile') and request.user.profile.level == '1' and genero_filtro == 'LICITACOES_E_CONTRATOS'):
            processos_query = processos_query.filter(
                genero='LICITACOES_E_CONTRATOS')
        elif (hasattr(request.user, 'profile') and request.user.profile.level == '2' and genero_filtro == 'LIQUIDACOES'):
            processos_query = processos_query.filter(genero='LIQUIDACOES')
        else:
            processos_query = processos_query.none()

    if especie_filtro != 'todas':
        processos_query = processos_query.filter(especie=especie_filtro)

    if status_analise_filtro != 'todas':
        processos_query = processos_query.filter(
            status_analise=status_analise_filtro)

    today = date.today()

    processos_pendentes_atrasados = processos_query.filter(
        status_monitoramento='PENDENTE',
        proxima_data_monitoramento__lt=today
    )
    for processo in processos_pendentes_atrasados:
        processo.status_monitoramento = 'ATRASADO'
        processo.save()

    processos_concluidos_a_reabrir = processos_query.filter(
        status_monitoramento='CONCLUIDO',
        proxima_data_monitoramento__isnull=False,
        proxima_data_monitoramento__lte=today
    )
    for processo in processos_concluidos_a_reabrir:
        processo.status_monitoramento = 'PENDENTE'
        processo.save()

    processos_query = processos_query.order_by('-data_saida')

    processos_data = []
    for processo in processos_query:
        prazo_obj = calcular_prazo(processo.data_entrada, processo.prioridade)
        if prazo_obj:
            dias_restantes = (prazo_obj - date.today()).days
            processo.prazo_formatado = f"{dias_restantes} dia(s) {'atrasado' if dias_restantes < 0 else 'restante(s)'}"
        else:
            processo.prazo_formatado = "-"

        p_dict = {
            'id': processo.id,
            'numero_processo': processo.numero_processo,
            'volume': processo.volume or '',
            'secretaria': processo.secretaria,
            'data_entrada': processo.data_entrada,
            'hora_entrada': processo.hora_entrada,
            'data_saida': processo.data_saida,
            'hora_saida': processo.hora_saida,
            'destino': processo.destino or '',
            'genero': processo.genero,
            'especie': processo.especie,
            'objeto': processo.objeto,
            'contratada': processo.contratada or '',
            'recorrente': processo.recorrente,
            'prioridade': processo.prioridade,
            'tecnico': processo.tecnico or '',
            'numero_despacho': processo.numero_despacho or '',
            'observacao': processo.observacao or '',
            'prazo_monitoramento_display': processo.get_prazo_monitoramento_display(),
            'proxima_data_monitoramento_formatada': processo.proxima_data_monitoramento.strftime('%d/%m/%Y') if processo.proxima_data_monitoramento else '',
            'status_monitoramento_display': processo.get_status_monitoramento_display(),
            'status_monitoramento_raw': processo.status_monitoramento,
            # NEW FIELDS
            'valor': processo.valor or '',
            'periodo': processo.periodo or '',
            'status_analise': processo.status_analise,  # Pass the raw value
            # Pass the display value
            'status_analise_display': processo.get_status_analise_display()
        }
        processos_data.append(p_dict)

    all_generos = Processo.objects.values_list(
        'genero', flat=True).distinct().order_by('genero')
    all_generos = [
        g for g in all_generos if can_access_genero(request.user, g)]

    all_especies = []
    if genero_filtro != 'todas':
        if can_access_genero(request.user, genero_filtro):
            all_especies = Processo.objects.filter(genero=genero_filtro).values_list(
                'especie', flat=True).distinct().order_by('especie')
    else:
        allowed_generos_for_all_species = [g for g in Processo.objects.values_list(
            'genero', flat=True).distinct() if can_access_genero(request.user, g)]
        all_especies = Processo.objects.filter(genero__in=allowed_generos_for_all_species).values_list(
            'especie', flat=True).distinct().order_by('especie')

    # Get all possible status_analise choices for the filter
    all_status_analise = [
        {'value': choice[0], 'label': choice[1]}
        for choice in Processo.STATUS_ANALISE_CHOICES
    ]

    # Pre-calculate flags for template logic
    can_edit = request.user.is_superuser or \
        (hasattr(request.user, 'profile')
         and request.user.profile.level in ['1', '2', '3'])

    can_delete = request.user.is_superuser

    can_mark_saida = request.user.is_superuser or \
        (hasattr(request.user, 'profile')
         and request.user.profile.level in ['0', '3'])

    return render(request, 'finalizados.html', {
        'processos': processos_data,
        'prioridade_filtro': prioridade_filtro,
        'termo_pesquisa': termo_pesquisa,
        'data_inicial_filtro': data_inicial_filtro,
        'data_final_filtro': data_final_filtro,
        'status_monitoramento_filtro': status_monitoramento_filtro,
        'especie_filtro': especie_filtro,
        'genero_filtro': genero_filtro,
        'status_analise_filtro': status_analise_filtro,
        'all_generos': all_generos,
        'all_especies': all_especies,
        'all_status_analise': all_status_analise,
        'is_admin': request.user.is_superuser,
        'user_level': request.user.profile.level if hasattr(request.user, 'profile') else None,
        'can_edit': can_edit,
        'can_delete': can_delete,
        'can_mark_saida': can_mark_saida,
    })


@login_required
@user_passes_test(lambda u: u.is_superuser or (hasattr(u, 'profile') and u.profile.level in ['0', '3']))
def exportar_finalizados_excel(request):
    data_inicial_str = request.GET.get('data_inicial')
    data_final_str = request.GET.get('data_final')
    prioridade = request.GET.get('prioridade')
    termo_pesquisa = request.GET.get('termo')
    status_monitoramento = request.GET.get('status_monitoramento', 'todas')
    especie = request.GET.get('especie', 'todas')
    genero = request.GET.get('genero', 'todas')
    status_analise = request.GET.get('status_analise', 'todas')

    if not data_inicial_str or not data_final_str:
        return HttpResponse('{"success": false, "message": "Por favor, selecione uma Data Inicial e uma Data Final para exportar os processos por período."}',
                            content_type='application/json', status=400)

    try:
        data_inicial = datetime.strptime(data_inicial_str, '%Y-%m-%d').date()
        data_final = datetime.strptime(data_final_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse('{"success": false, "message": "Formato de data inválido. Use AAAA-MM-DD."}',
                            content_type='application/json', status=400)

    processes = Processo.objects.filter(
        data_saida__range=[data_inicial, data_final]
    ).order_by('data_saida')

    processes = filter_processes_by_user_level(request.user, processes)

    if prioridade and prioridade != 'todas':
        processes = processes.filter(prioridade=prioridade)
    if termo_pesquisa:
        processes = processes.filter(
            Q(numero_processo__icontains=termo_pesquisa) |
            Q(secretaria__icontains=termo_pesquisa) |
            Q(destino__icontains=termo_pesquisa) |
            Q(genero__icontains=termo_pesquisa) |
            Q(especie__icontains=termo_pesquisa) |
            Q(objeto__icontains=termo_pesquisa) |
            Q(contratada__icontains=termo_pesquisa) |
            Q(tecnico__icontains=termo_pesquisa) |
            Q(observacao__icontains=termo_pesquisa) |
            Q(valor__icontains=termo_pesquisa) |
            Q(periodo__icontains=termo_pesquisa)
        )

    if status_monitoramento != 'todas':
        processes = processes.filter(status_monitoramento=status_monitoramento)

    if genero and genero != 'todas':
        if can_access_genero(request.user, genero):
            processes = processes.filter(genero=genero)
        else:
            processes = processes.none()

    if especie and especie != 'todas':
        processes = processes.filter(especie=especie)

    if status_analise != 'todas':
        processes = processes.filter(status_analise=status_analise)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Processos Finalizados"

    headers = [
        "N° Processo", "Volume", "Secretaria", "Data Entrada", "Hora Entrada",
        "Data Saída", "Hora Saída", "Destino", "Gênero", "Espécie", "Objeto",
        "Contratada", "Recorrente", "Prioridade", "Técnico", "N° Despacho",
        "Observação", "Valor"
    ]
    sheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(
        start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for process in processes:
        row_data = [
            process.numero_processo,
            process.volume,
            process.secretaria,
            process.data_entrada.strftime(
                '%Y-%m-%d') if process.data_entrada else '',
            process.hora_entrada.strftime(
                '%H:%M') if process.hora_entrada else '',
            process.data_saida.strftime(
                '%Y-%m-%d') if process.data_saida else '',
            # Formatacao Excel
            process.hora_saida.strftime('%H:%M') if process.hora_saida else '',
            process.destino,
            process.genero,
            process.especie,
            process.objeto,
            process.contratada,
            process.recorrente,
            process.prioridade,
            process.tecnico,
            process.numero_despacho,
            process.observacao,
            process.valor,

        ]
        sheet.append(row_data)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            if cell.column_letter in ['K', 'O', 'P', 'Q', 'R', 'U', 'V', 'W']:
                cell.alignment = Alignment(wrapText=True, vertical='top')

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = (max_length + 2)
        if column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']:
            sheet.column_dimensions[column_letter].width = min(
                adjusted_width, 100)
        else:
            sheet.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=processos_finalizados_{data_inicial_str}_a_{data_final_str}.xlsx'
    workbook.save(response)

    return response


@login_required
def get_process_by_number(request, numero_processo):
    try:
        processo = Processo.objects.filter(numero_processo=numero_processo).order_by(
            '-data_entrada', '-hora_entrada').first()

        if not processo:
            return JsonResponse({'message': 'Processo não encontrado'}, status=404)

        if not can_access_genero(request.user, processo.genero):
            return JsonResponse({"message": "Você não tem permissão para visualizar detalhes deste processo."}, status=403)

        data = {
            'numero_processo': processo.numero_processo,
            'volume': processo.volume,
            'secretaria': processo.secretaria,
            'data_entrada': processo.data_entrada.strftime('%Y-%m-%d') if processo.data_entrada else None,
            'hora_entrada': processo.hora_entrada.strftime('%H:%M') if processo.hora_entrada else None,
            'data_saida': processo.data_saida.strftime('%Y-%m-%d') if processo.data_saida else None,
            # Format Time object for JSON
            'hora_saida': processo.hora_saida.strftime('%H:%M') if processo.hora_saida else None,
            'destino': processo.destino,
            'genero': processo.genero,
            'especie': processo.especie,
            'objeto': processo.objeto,
            'contratada': processo.contratada,
            'recorrente': processo.recorrente,
            'prioridade': processo.prioridade,
            'tecnico': processo.tecnico,
            'data_analise': processo.data_analise.strftime('%Y-%m-%d') if processo.data_analise else None,
            'numero_despacho': processo.numero_despacho,
            'observacao': processo.observacao,
            'prazo_monitoramento': processo.prazo_monitoramento,
            'proxima_data_monitoramento': processo.proxima_data_monitoramento.strftime('%Y-%m-%d') if processo.proxima_data_monitoramento else None,
            'status_monitoramento': processo.status_monitoramento,
            'valor': processo.valor,
            'periodo': processo.periodo,
            'status_analise': processo.status_analise,
        }
        return JsonResponse(data)
    except Exception as e:
        print(f"Erro inesperado em get_process_by_number: {e}")
        return JsonResponse({'message': 'Erro interno do servidor'}, status=500)


@csrf_exempt
@login_required
def marcar_saida_processo(request, process_id):
    if request.method == 'POST':
        try:
            processo = get_object_or_404(Processo, id=process_id)

            if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.level in ['0', '3'])):
                return JsonResponse({"success": False, "message": "Você não tem permissão para marcar a saída deste processo."}, status=403)

            if processo.data_saida:
                return JsonResponse({"success": False, "message": "A data de saída já está preenchida para este processo."}, status=400)

            now_local = timezone.localtime(timezone.now())

            processo.data_saida = now_local.date()
            processo.hora_saida = now_local.time()

            ProcessHistory.objects.create(
                process=processo,
                field_name='data_saida',
                old_value=None,
                new_value=now_local.date().strftime('%Y-%m-%d'),
                changed_by=request.user
            )
            ProcessHistory.objects.create(
                process=processo,
                field_name='hora_saida',
                old_value=None,
                new_value=now_local.time().strftime('%H:%M:%S'),
                changed_by=request.user
            )

            processo.save()

            return JsonResponse({
                "success": True,
                "message": "Data e hora de saída marcadas com sucesso!",
                "data_saida": now_local.strftime('%Y-%m-%d'),
                "hora_saida": now_local.strftime('%H:%M')
            })

        except Processo.DoesNotExist:
            return JsonResponse({"success": False, "message": "Processo não encontrado."}, status=404)
        except Exception as e:
            print(f"Erro ao marcar saída do processo: {e}")
            return JsonResponse({"success": False, "message": str(e)}, status=500)
    return JsonResponse({"success": False, "message": "Método não permitido."}, status=405)


def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('listar_processos')
            else:
                return render(request, 'registration/login.html', {'form': form, 'error': 'Nome de usuário ou senha inválidos.'})
    else:
        form = AuthenticationForm()
    return render(request, 'registration/login.html', {'form': form})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            return redirect('manage_users')
        else:
            return render(request, 'registration/register.html', {'form': form})
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/register.html', {'form': form})


@login_required
@user_passes_test(lambda u: u.is_superuser or u.is_staff)
def manage_users(request):
    users = User.objects.all().order_by('username')
    users = users.select_related('profile')

    user_levels = Profile.USER_LEVEL_CHOICES

    return render(request, 'admin/manage_users.html', {'users': users, 'user_levels': user_levels})


@login_required
@user_passes_test(lambda u: u.is_superuser)
def update_user_level(request, user_id):
    if request.method == 'POST':
        user_to_update = get_object_or_404(User, id=user_id)
        if user_to_update.id == request.user.id or user_to_update.is_superuser:
            return redirect('manage_users')

        new_level = request.POST.get('level')
        if new_level and new_level in [choice[0] for choice in Profile.USER_LEVEL_CHOICES]:
            profile, created = Profile.objects.get_or_create(
                user=user_to_update)
            profile.level = new_level
            profile.save()
        return redirect('manage_users')
    return redirect('manage_users')


@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_user(request, user_id):
    if request.method == 'POST':
        user_to_delete = get_object_or_404(User, id=user_id)
        if user_to_delete.id == request.user.id:
            return redirect('manage_users')
        user_to_delete.delete()
        return redirect('manage_users')
    return redirect('manage_users')


@login_required
def get_especies_by_genero(request):
    genero = request.GET.get('genero')
    if genero and can_access_genero(request.user, genero):
        especies = Processo.objects.filter(genero=genero).values_list(
            'especie', flat=True).distinct().order_by('especie')
        return JsonResponse({'especies': list(especies)})
    return JsonResponse({'especies': []})


@login_required
def get_all_especies(request):
    allowed_generos = [g for g in Processo.objects.values_list(
        'genero', flat=True).distinct() if can_access_genero(request.user, g)]
    especies = Processo.objects.filter(genero__in=allowed_generos).values_list(
        'especie', flat=True).distinct().order_by('especie')
    return JsonResponse({'especies': list(especies)})
